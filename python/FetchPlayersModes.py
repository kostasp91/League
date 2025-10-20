from playwright.sync_api import sync_playwright
from openpyxl import Workbook, load_workbook
import os

# === CONFIGURATION ===
MODE = "update"  # "full" → rebuild from scratch, "update" → append new weeks
OUTPUT_FILE = "players_data_dunkest.xlsx"
WEEKS = 5     # total number of weeks to fetch
ROUNDS = 2    # number of rounds per week


# === HELPER FUNCTIONS ===
def parse_float(value):
    """Convert string to float, handling commas, percentages, and minus signs."""
    if value is None:
        return 0.0
    value = value.strip().replace(",", ".").replace("\u2212", "-")
    if value.endswith("%"):
        value = value[:-1]
    try:
        return float(value)
    except ValueError:
        return 0.0


def parse_row(cols, week, rnd):
    """Extract player data from table row cells, skipping the # column safely."""
    # Some rows might be incomplete (ads, totals, etc.)
    #if len(cols) < 31:
    #    return None

    # Drop the "#" cell (the first one)
    #cols = cols[1:]

    try:
        return [
            cols[0].inner_text().strip(),   # Player
            cols[1].inner_text().strip(),   # Pos
            cols[2].inner_text().strip(),   # Team
            parse_float(cols[3].inner_text()),   # FPT
            parse_float(cols[4].inner_text()),   # CR
            parse_float(cols[5].inner_text()),   # PLUS
            parse_float(cols[6].inner_text()),   # GP
            parse_float(cols[7].inner_text()),   # MIN
            parse_float(cols[8].inner_text()),   # ST
            parse_float(cols[9].inner_text()),   # PTS
            parse_float(cols[10].inner_text()),  # REB
            parse_float(cols[11].inner_text()),  # AST
            parse_float(cols[12].inner_text()),  # STL
            parse_float(cols[13].inner_text()),  # BLK
            parse_float(cols[14].inner_text()),  # BA
            parse_float(cols[15].inner_text()),  # FGM
            parse_float(cols[16].inner_text()),  # FGA
            parse_float(cols[17].inner_text()),  # FG%
            parse_float(cols[18].inner_text()),  # 3PM
            parse_float(cols[19].inner_text()),  # 3PA
            parse_float(cols[20].inner_text()),  # 3P%
            parse_float(cols[21].inner_text()),  # FTM
            parse_float(cols[22].inner_text()),  # FTA
            parse_float(cols[23].inner_text()),  # FT%
            parse_float(cols[24].inner_text()),  # OREB
            parse_float(cols[25].inner_text()),  # DREB
            parse_float(cols[26].inner_text()),  # TOV
            parse_float(cols[27].inner_text()),  # PF
            parse_float(cols[28].inner_text()),  # FD
            parse_float(cols[29].inner_text()),  # +/-
            week,
            rnd,
        ]
    except Exception:
        return None


# === MAIN SCRAPER FUNCTION ===
def scrape_data():
    # === MODE HANDLING ===
    if MODE.lower() == "full" or not os.path.exists(OUTPUT_FILE):
        print("Running in FULL mode — creating new file.")
        wb = Workbook()
        # remove default empty sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        start_week = 1
    elif MODE.lower() == "update":
        print("Running in UPDATE mode — adding missing weeks.")
        wb = load_workbook(OUTPUT_FILE)
        existing_sheets = wb.sheetnames
        last_week_num = 0
        for s in existing_sheets:
            if s.startswith("Week "):
                try:
                    week_num = int(s.split(" ")[1])
                    last_week_num = max(last_week_num, week_num)
                except Exception:
                    pass
        start_week = last_week_num + 1
    else:
        raise ValueError(f"Invalid MODE: {MODE}. Use 'full' or 'update'.")

    print(f"Starting from Week {start_week}")

    # === SCRAPE LOOP ===
    with sync_playwright() as p:
        browser = p.firefox.launch(headless=True)
        page = browser.new_page()

        for week in range(start_week, WEEKS + 1):
            ws = wb.create_sheet(title=f"Week {week}")

            # headers (without the # column)
            ws.append([
                "Player", "Pos", "Team", "FPT", "CR", "PLUS", "GP", "MIN", "ST", "PTS",
                "REB", "AST", "STL", "BLK", "BA", "FGM", "FGA", "FG%", "3PM", "3PA",
                "3P%", "FTM", "FTA", "FT%", "OREB", "DREB", "TOV", "PF", "FD", "+/-",
                "Week", "Round"
            ])

            for rnd in range(1, ROUNDS + 1):
                print(f"Fetching Week {week}, Round {rnd}...")
                url = (
                    f"https://www.dunkest.com/en/euroleague/stats/players/table?"
                    f"season_id=23&mode=dunkest&stats_type=tot&weeks[]={week}&rounds[]={rnd}"
                    f"&teams[]=32&teams[]=33&teams[]=34&teams[]=35&teams[]=36&teams[]=37"
                    f"&teams[]=38&teams[]=39&teams[]=40&teams[]=41&teams[]=42&teams[]=43"
                    f"&teams[]=44&teams[]=45&teams[]=46&teams[]=47&teams[]=48&teams[]=56"
                    f"&teams[]=60&teams[]=75&positions[]=1&positions[]=2&positions[]=3"
                    f"&player_search=&min_cr=4&max_cr=35&sort_by=pdk&sort_order=desc"
                    f"&iframe=yes&noadv=yes"
                )

                page.goto(url)
                page.wait_for_selector("table tbody tr")

                # scrape table pages
                while True:
                    rows = page.query_selector_all("table tbody tr")
                    for row in rows:
                        cols = row.query_selector_all("td")
                        parsed = parse_row(cols, week, rnd)
                        if parsed:
                            ws.append(parsed)

                    # next page check
                    next_btn = page.query_selector("a[href='']:has-text('»')")
                    if not next_btn or "disabled" in (next_btn.get_attribute("class") or ""):
                        break
                    next_btn.click()
                    page.wait_for_timeout(1000)

        browser.close()

    # === SAVE RESULTS ===
    wb.save(OUTPUT_FILE)
    print(f" Done! Data saved cleanly to {OUTPUT_FILE}.")


# === ENTRY POINT ===
if __name__ == "__main__":
    scrape_data()
